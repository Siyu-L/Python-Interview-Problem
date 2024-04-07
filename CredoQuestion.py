#Siyu Li
import openpyxl
import matplotlib.pyplot as plt
import random
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#Asks user to choose 2 columns to plot, using index column as x-axis and the 2 columns as y-axis
#param dataTable: data table from which rows are chosen from
#param numRows: number of rows in dataTable
#param numCols: number of columns in dataTable

def plot_columns(dataTable, numRows, numCols):

	#Continuously asks user until valid columns are entered
	#Column is valid if it's between 0 and column number - 2 and if the second column is not the same as the first
	#note: range of columns is between 0 and numCols - 2 because an additional column is added as the index column

	while True:
		firstColumn = int(input("Enter 1st column to graph (0-" + str(numCols-2) + "): "))
		if firstColumn >=0 and firstColumn <= numCols - 2:
			break
		else:
			print("Invalid column.")
	while True:
		secondColumn = int(input("Enter 2nd column to graph (0-" + str(numCols-2) + "): "))
		if secondColumn >=0 and secondColumn <= numCols - 2 and secondColumn != firstColumn:
			break
		else:
			print("Invalid column.")
	
	#create temporary arrays for x and y values
	yPoints1 = [0 for x in range(numRows)]
	yPoints2 = [0 for x in range(numRows)]
	xPoints = [0 for x in range(numRows)]	
	
	#fill yPoint arrays with values from two selected columns
	#fill xPoint array with values from first(index) column
	i = 0
	while(i < numRows):
		yPoints1[i] = dataTable[i][firstColumn+1]
		yPoints2[i] = dataTable[i][secondColumn+1]
		#note: index of a chosen column is num + 1 because first column is used to index rows
		
		xPoints[i] = dataTable[i][0]
		i = i+1
	xPoints.pop(0) #remove word "index"
	yPoints1.pop(0) #remove index of 1st column
	yPoints2.pop(0) #remove index of 2nd column

	#plot each column
	plt.plot(xPoints, yPoints1, 'bo')
	plt.plot(xPoints, yPoints2, 'bo')
	
	#save plot
	plt.savefig('sample.png')




cols = int(input("Enter number of columns: ")) + 1
#number of columns = number of columns of random values + 1 column for index
#if number of columns is 20, cols = 21


rows = int(input("Enter number of rows: ")) + 1
#number of rows = number of rows of random values + 1 row for naming
#if number of rows is 200, rows = 201

data = [[0 for i in range(cols)] for j in range(rows)]
#2D array storing the data

i = 0 #iterating rows
while(i < rows):
	j = 0 #iterating columns
	while(j < cols):
		if i ==0:
			data[i][j] = "col " + str(j-1) #first row naming
		elif j == 0:
			data[i][j] = int(i - 1) #numbering index column
		else:
			data[i][j] = random.random() #generates random float in range [0.0,1.0)
		j = j+1
	i = i+1

data[0][0] = "index"

try:
	#input data into excel spreadsheet
	wb = openpyxl.Workbook()
	ws = wb.active

	ws.title = "Sample"

	for x in range(rows):
		ws.append(data[x])

	#saved as "sample.xlsx"
	wb.save(filename = "sample.xlsx")

	#open and read the excel file
	new_wb = load_workbook("sample.xlsx")
	new_ws = wb.active

	#create and read data into new array
	new_data = [[0 for x in range(cols)] for y in range(rows)]
	for i in range(1, rows+1):
		for j in range(1, cols+1):
			temp_char = get_column_letter(j)
			new_data[i-1][j-1] = ws[temp_char + str(i)].value

except Exception as e:
	print(e)
	print("Error: Trying to modify open excel file")
#print out data read from array
for r in new_data:
	print(r)


#call function to plot columns
plot_columns(new_data, rows, cols)
